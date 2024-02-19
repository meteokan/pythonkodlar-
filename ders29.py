import tkinter as tk
from openpyxl import load_workbook, Workbook
def kayit_yap():
    ad=entry_ad.get()
    soyad=entry_soyad.get()
    yas=entry_yas.get()
    try:
        dosya=load_workbook("kayitlar.xlsx")
        sheet=dosya.active
        son_satir = sheet.max_row
    except FileNotFoundError:
        dosya = Workbook()
        sheet=dosya.active
        sheet.append(["Ad","Soyad","Yaş"])
        dosya.save("kayitlar.xlsx")
    yeni_satir=[ad,soyad,yas]
    sheet.append(yeni_satir)
    dosya.save("kayitlar.xlsx")
    label_sonuc.config(text="kayıt başarılı")

arayuz=tk.Tk()
arayuz.title("Kayıt Ekranı")

label_ad=tk.Label(arayuz, text="Ad:")
label_ad.grid(row=0, column=0)
entry_ad=tk.Entry(arayuz)
entry_ad.grid(row=0, column=1)

label_soyad=tk.Label(arayuz, text="Soyad:")
label_soyad.grid(row=1, column=0)
entry_soyad=tk.Entry(arayuz)
entry_soyad.grid(row=1, column=1)

label_yas=tk.Label(arayuz, text="Yaş:")
label_yas.grid(row=2, column=0)
entry_yas=tk.Entry(arayuz)
entry_yas.grid(row=2, column=1)

button_kayit=tk.Button(arayuz, text="Kayıt", command=kayit_yap )
button_kayit.grid(row=3, columnspan=2)

label_sonuc=tk.Label(arayuz, text="")
label_sonuc.grid(row=4, column=0)

arayuz.mainloop()


