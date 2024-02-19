import openpyxl
from openpyxl.styles import Font
dosya=openpyxl.load_workbook("yeni_excel.xlsx")
sayfa=dosya.active
print(dosya.sheetnames)
print(dosya.sheetnames[0])
# print(sayfa)
dosya.active=0
# print(sayfa)
sayfa.cell(row=1, column=4, value="BoyKilo_İndeksi")
sayfa.cell(row=1, column=5, value="Durum")
satir_sayisi=sayfa.max_row
print(satir_sayisi)
for satir in range(2,satir_sayisi+1):
    boy=sayfa.cell(satir,2).value
    kilo=sayfa.cell(satir,3).value
    boykilo_indeksi=round(kilo/(boy**2))
    sayfa.cell(row=satir, column=4, value=boykilo_indeksi)
    if(boykilo_indeksi<18):
        durum="Zayıf"
        sayfa.cell(row=satir, column=5, value=durum)
        sayfa.cell(satir,5).font=Font(name="Times New Roman",color="ff0000")
    elif(boykilo_indeksi<25):
        durum = "Normal"
        sayfa.cell(row=satir, column=5, value=durum)
        sayfa.cell(satir, 5).font = Font(name="Times New Roman", color="00ff00")
    elif(boykilo_indeksi<30):
        durum = "Kilo"
        sayfa.cell(row=satir, column=5, value=durum)
    else:
        durum = "hata"
        sayfa.cell(row=satir, column=5, value=durum)
dosya.save("yeni_excel.xlsx")

