import openpyxl
dosya=openpyxl.load_workbook("yeni_excel.xlsx")
sayfa=dosya["Deneme"]
satir_sayisi=sayfa.max_row
sutun_sayisi=sayfa.max_column
data=[]
for i in range(2,satir_sayisi+1):
    satir=[]
    for j in range(1,sutun_sayisi+1):
       satir.append(sayfa.cell(i,j).value)
    data.append(satir)
print(data)
for i in range(len(data)):
    print(data[i])
print(data[0][0])
print(data[1][2]-10)
print(f"{data[1][0]} diyet yaptı ve {data[1][2]-10} kiloya düştü")