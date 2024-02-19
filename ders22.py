import openpyxl
dosya=openpyxl.load_workbook("dersexcel.xlsx")
sayfa=dosya["uyeler"]
sayfa.cell(row=1,column=4,value="deger")
satir_sayisi=sayfa.max_row
for satir in range(2,satir_sayisi+1):
    ucret=sayfa.cell(satir,3).value
    deger=(ucret*2)
    sayfa.cell(row=satir,column=4,value=deger)
    sayfa.cell(satir,4)
dosya.save("dersexcel.xlsx")
