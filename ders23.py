import openpyxl
from openpyxl import Workbook
dosya=Workbook()
sayfa=dosya.active
sayfa.title="Deneme"
sayfa["A1"].value="İsimler"
sayfa["B1"].value="Boy"
sayfa["C1"].value="Kilo"
dosya.create_sheet("Profil")
dosya.create_sheet("Ücret")
deger=dosya["Profil"]["B2"].value="İşlem tamam."
dosya.active=1
print(dosya.sheetnames)
dosya.save("yeni_excel.xlsx")
