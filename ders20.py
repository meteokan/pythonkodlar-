import openpyxl
dosya=openpyxl.load_workbook("dersexcel.xlsx")
print(dosya.sheetnames)
# 1. yöntem
sayfa= dosya["uyeler"]
deger=sayfa["C2"].value
print(deger)
# 2. yöntem
deger=dosya["uyeler"]["c2"].value
print(deger)
# 3. yöntem
sayfa=dosya["uyeler"]
deger1=sayfa.cell(2,3).value
print("notu: ",deger1)


