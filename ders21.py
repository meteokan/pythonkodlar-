from openpyxl import load_workbook
dosya=load_workbook("dersexcel.xlsx")
sayfa=dosya.active
print(dosya.sheetnames)
print(sayfa)
sayfa=dosya["hisseler"]
print(sayfa)
print(sayfa.cell(3,2).value)
for satir in sayfa["A1":"C4"]:
    for hucre in satir:
        print("->",str(hucre.value),"<-",end="")
    print()
sayfa["a4"]="özerler"
sayfa["A5"]="meteler"
sayfa["B5"]=100
for satir in sayfa["A1":"C5"]:
    for hucre in satir:
        print("->",str(hucre.value),"<-",end="")
    print()
dosya.save("dersexcel.xlsx")





