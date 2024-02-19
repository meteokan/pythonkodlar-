from openpyxl import load_workbook
Dosya=load_workbook("dosya_adi.xlsx")
print(Dosya.sheetnames)
sayfa=Dosya["Yeniden_Adlandırılan_Sayfa"]
sayfa.delete_cols(4,amount=1)
sayfa.delete_rows(3,amount=1)
sayfa.insert_cols(4,amount=4)

Dosya.save("dosya_adi1.xlsx")